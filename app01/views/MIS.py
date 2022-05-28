import xlrd
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt
from app01 import models
from openpyxl import load_workbook
from django import forms


def MIS(request):
    info = request.session.get("info")
    if not info:
        return redirect('/login/')
    if request.method == "GET":
        return render(request, 'html/MIS/MIS.html')
    return render(request, 'html/MIS/MIS.html')


@csrf_exempt
def mis_test(request):
    print(request.POST)
    data_dict = {'status': True, 'data': [11, 22, 33, 44]}
    return JsonResponse(data_dict)


@csrf_exempt
def mis_chart(request):
    info = request.session.get("info")
    if not info:
        return redirect('/login/')
    return render(request, 'html/MIS/mis_chart.html')


def mis_chart_bar(request):
    if request.method == "GET":
        # 构造柱状图
        legend = ['chenzhewei', 'hll']
        series_list = [
            {
                "name": "chenzhewei",
                "type": 'bar',
                "data": [1, 2, 3, 4, 5, 6]
            },
            {
                "name": "hll",
                "type": 'bar',
                "data": [2, 3, 4, 5, 6, 7]
            }
        ]
        x_axis = ['1月', '2月', '3月', '4月', '5月', '6月']
        result = {
            "status": True,
            "data": {
                "legend": legend,
                "series_list": series_list,
                "x_axis": x_axis,
            }
        }
        return JsonResponse(result)


@csrf_exempt
def mis_chart_line(request):
    if request.method == "GET":
        queryset = models.Mis.objects.all()
        legend = ["3MIS", "6MIS", "12MIS"]
        x_axis = []
        data_3mis = []
        data_6mis = []
        data_12mis = []
        title = queryset[0].name_of_parts + "MIS趋势图"
        for form in queryset:
            # 将x轴的数据写入到数组内
            x_axis.append(form.date)
            data_3mis.append(float(form.mis_3))
            data_6mis.append(float(form.mis_6))
            data_12mis.append(float(form.mis_12))
        series_list = [
            {
                "name": "3MIS",
                "type": "line",
                "data": data_3mis,
                "stack": "Total",
            },
            {
                "name": "6MIS",
                "type": "line",
                "data": data_6mis,
                "stack": "Total",
            },
            {
                "name": "12MIS",
                "type": "line",
                "data": data_12mis,
                "stack": "Total",
            },
        ]
        result = {
            "status": True,
            "data": {
                "title": title,
                "legend": legend,
                "series_list": series_list,
                "x_axis": x_axis,
            }
        }
        return JsonResponse(result)


class mis_form(forms.ModelForm):
    class Meta:
        model = models.Mis
        fields = "__all__"


def mis_upload(request):
    """基于Excel的文件上传"""
    # 1. 获取用户上传的文件对象
    form = mis_form(data=request.POST, files=request.FILES)
    file_object = request.FILES.get("filename", None)
    if not file_object:
        return HttpResponse("没有文件可供上传")
    # 2. 将文件对象传递给openpyxl，由openpyxl来进行读取
    wb = load_workbook(file_object)
    # 3. 将读取出来的第一个表格传入到sheet里面
    sheet = wb.worksheets[0]
    dict_list = []
    for row in sheet.iter_rows(min_row=2):
        for item in range(len(row)):
            text = str(row[item].value)
            dict_list.append(text)
        models.Mis.objects.create(name_of_parts=dict_list[0], mis_3=dict_list[1], mis_6=dict_list[2],
                                  mis_12=dict_list[3], date=dict_list[4])
        dict_list.clear()
        # if form.is_valid():
        #     form.save()
    return redirect('/MIS/chart/')
